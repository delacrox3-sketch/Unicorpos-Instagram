#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Gera o calendario.json a partir da planilha UNICORPOS_Instagram_Calendario30d_v3.xlsx.

Roda LOCALMENTE (nao no GitHub Actions). O resultado, calendario.json, e' o que o
robo le todo dia. Depois de rodar, confira o JSON e faca commit.

    python3 ferramentas/gerar_calendario.py

Por que existe: o robo nao le xlsx. O JSON e' auditavel, versionado em git e voce
consegue corrigir uma legenda no dia sem abrir o Excel.
"""

import json
import os
import re
import sys

try:
    import openpyxl
except ImportError:
    sys.exit("Falta o openpyxl. Rode: pip install openpyxl")

AQUI = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

PLANILHA_PADRAO = os.path.join(
    os.path.expanduser("~"),
    "Claude", "Projects", "UNICORPOS", "deliverables", "documents",
    "UNICORPOS_Instagram_Calendario30d_v3.xlsx",
)

# ---------------------------------------------------------------------------
# Categorias. "odonto" nunca publica sozinho: vai por e-mail para o Leonardo.
# Vedacao do Codigo de Etica Odontologica (CFO) + exigencia de citar a RT.
# ---------------------------------------------------------------------------
CATEGORIA = {
    "F4_consultorio": "odonto",
    "08_odontologia": "odonto",
    "10_odonto_precos": "odonto",
    "17_pos_laser": "estetica",
    "18_primeira_fisio": "fisio",
    "19_biosseguranca": "institucional",
    "F9_porta_fisio": "fisio",
    "F6_fisio": "fisio",
    "11_pilates_rpg": "fisio",
    "16_quiropraxia": "fisio",
    "07_fisioterapia": "fisio",
    "F1_sala_estetica": "estetica",
    "F2_laser": "estetica",
    "F3_equipamentos": "estetica",
    "03_laser_face": "estetica",
    "04_laser_corpo": "estetica",
    "05_estetica_facial": "estetica",
    "06_massagens": "estetica",
}

# ---------------------------------------------------------------------------
# Texto alternativo, escrito olhando cada peca. A API aceita alt_text so em
# imagem de feed (nao em reels nem stories). Limite pratico: 1000 caracteres.
# ---------------------------------------------------------------------------
ALT = {
    "F7_recepcao": "Recepcao da UNICORPOS: balcao escuro com o logotipo dourado da clinica na parede, iluminacao embutida circular no teto e lustre de cristal. Texto sobreposto: Cuidar de voce e o nosso proposito.",
    "F10_escada": "Area de circulacao interna da clinica, com lustre de cristal e painel iluminado da UNICORPOS ao fundo. Texto sobreposto: Um espaco pensado para voce.",
    "F9_porta_fisio": "Painel externo dourado da unidade DF-130 com o logotipo UNICORPOS e a inscricao Fisioterapia e Quiropraxia. Texto sobreposto: Unidade DF-130, sentido Vale do Amanhecer, com estacionamento no local.",
    "F1_sala_estetica": "Sala de estetica da unidade DF-130, com maca clara, armarios de madeira, persianas e equipamentos ao fundo sob a placa Estetica. Texto sobreposto: Sala de estetica.",
    "F4_consultorio": "Consultorio odontologico da UNICORPOS com duas estacoes de trabalho, monitores e persianas, sob a placa Odontologia. Texto sobreposto: Consulta de avaliacao gratuita.",
    "03_laser_face": "Card dourado e creme da UNICORPOS com a tabela de valores por area de depilacao a laser facial.",
    "F2_laser": "Equipamento de depilacao a laser HAKON, em azul escuro, instalado na sala de atendimento da unidade DF-130. Texto sobreposto: Tecnologia HAKON.",
    "12_horarios": "Card da UNICORPOS com o horario de funcionamento: segunda a sexta das 8h as 12h e das 14h as 18h, sabado das 8h as 12h, domingo e feriados fechado.",
    "F6_fisio": "Sala de fisioterapia da UNICORPOS com maca de atendimento, plantas e escrivaninha junto a janela. Texto sobreposto: Atendimento individual.",
    "04_laser_corpo": "Card dourado e creme da UNICORPOS com a tabela de valores e pacotes de depilacao a laser corporal.",
    "F8_espera": "Sala de espera da UNICORPOS com poltronas escuras diante de uma parede envidracada com vista para a area externa, e lustre de cristal ao fundo. Texto sobreposto: Conforto de quem espera pouco.",
    "05_estetica_facial": "Card da UNICORPOS sobre limpeza de pele, peelings e tecnologias faciais na unidade da DF-130.",
    "F3_equipamentos": "Equipamentos de estetica corporal da UNICORPOS, incluindo aparelho de ultrassom focado, em sala com parede texturizada. Texto sobreposto: Equipamento para cada protocolo.",
    "14_como_funciona": "Card da UNICORPOS explicando os tres passos da primeira visita: avaliacao, plano e valor, e consentimento. Rodape informa que nenhum resultado e garantido.",
    "F5_ambiente": "Sala de atendimento da UNICORPOS com equipamentos de estetica, maca e bancada clara junto a janela. Texto sobreposto: Conforto e biosseguranca.",
    "11_pilates_rpg": "Card da UNICORPOS com os valores de Pilates clinico individual, em dupla e em grupo, alem de liberacao miofascial e massoterapia.",
    "16_quiropraxia": "Card da UNICORPOS explicando o que e quiropraxia e para quais casos ela e indicada.",
    "13_unidades": "Card da UNICORPOS com o endereco das duas unidades: odontologia no Setor Tradicional e fisioterapia e estetica na DF-130.",
    "08_odontologia": "Card da UNICORPOS sobre a consulta de avaliacao gratuita em odontologia, com plano de tratamento e orcamento por escrito.",
    "09_agendar": "Card da UNICORPOS convidando para agendar atendimento pelo WhatsApp, com hora marcada de segunda a sabado.",
    "01_capa": "Card de apresentacao da UNICORPOS, clinica de odontologia, estetica e fisioterapia em Planaltina, Distrito Federal.",
    "02_nova_unidade": "Card anunciando a nova unidade da UNICORPOS na DF-130, KM 0, sentido Vale do Amanhecer, com estacionamento no local.",
    "06_massagens": "Card da UNICORPOS sobre massagens e terapias corporais: relaxante, modeladora, drenagem linfatica e liberacao miofascial.",
    "07_fisioterapia": "Card da UNICORPOS com os valores de fisioterapia ortopedica, reabilitacao de coluna, RPG e atendimento pos-operatorio.",
    "15_lgpd": "Card da UNICORPOS explicando que os dados dos pacientes sao usados apenas para agendamento e atendimento, conforme a Lei 13.709/2018.",
    "10_odonto_precos": "Card da UNICORPOS sobre estetica dental: clareamento, facetas, lentes de contato, proteses, coroas e implantes, com avaliacao gratuita.",
    "17_pos_laser": "Card da UNICORPOS com cuidados apos a sessao de depilacao a laser: evitar sol, calor e exercicio nas primeiras 24 horas, usar protetor solar todo dia, e o aviso de que vermelhidao leve por algumas horas e esperada.",
    "18_primeira_fisio": "Card da UNICORPOS explicando os tres passos de quem nunca fez fisioterapia: avaliacao, plano de sessoes com valor informado, e atendimento individual com hora marcada.",
    "19_biosseguranca": "Card da UNICORPOS sobre a rotina de higienizacao: maca e aparelhos limpos a cada paciente, material de uso unico descartado na frente do paciente, e salas individuais climatizadas.",
}

# ---------------------------------------------------------------------------
# Legenda que nao existia na planilha. F10_escada tem imagem pronta em 1080x1080,
# entao vai ao feed como post normal em vez de virar story.
# ---------------------------------------------------------------------------
LEGENDAS_EXTRA = {
    "17_pos_laser": (
        "Fez laser? Os dois dias seguintes contam tanto quanto a sessão. ☀️\n\n"
        "Sem sol direto, sem água muito quente, sem sauna e sem treino pesado nas "
        "primeiras 24 horas. E protetor solar na área tratada todo dia, mesmo com "
        "tempo fechado.\n\n"
        "Vermelhidão leve por algumas horas é esperado. Se passar disso, chama a "
        "gente que a gente olha.\n\n"
        "WhatsApp (61) 99578-9867\n\n"
        "#unicorpos #depilacaoalaser #planaltinadf #cuidadoscomapele"
    ),
    "18_primeira_fisio": (
        "Nunca fez fisioterapia e não sabe como começa? É assim. 🤍\n\n"
        "Primeiro uma avaliação: conversa e testes de movimento para entender sua "
        "queixa. Depois você recebe o plano — quantas sessões, com que frequência e "
        "quanto custa, antes de começar qualquer coisa.\n\n"
        "O atendimento é individual, com hora marcada. Uma pessoa por vez.\n\n"
        "Responsável Técnico: Mario Sergio Fernandes de Lima — CREFITO-11 nº 442563-F\n"
        "WhatsApp (61) 99578-9867\n\n"
        "#unicorpos #fisioterapia #planaltinadf #valedoamanhecer"
    ),
    "19_biosseguranca": (
        "Tem uma parte do atendimento que você não vê — e é bom que seja assim. 🧼\n\n"
        "Entre um paciente e outro, maca, aparelhos e superfícies são higienizados. "
        "O que é de uso único é aberto e descartado na sua frente. As salas são "
        "individuais, climatizadas e ventiladas.\n\n"
        "Biossegurança não é diferencial. É obrigação.\n\n"
        "WhatsApp (61) 99578-9867\n\n"
        "#unicorpos #planaltinadf #biosseguranca #clinicaespecializada"
    ),
    "F10_escada": (
        "Um espaco pensado para voce. 🤍\n\n"
        "Da recepcao a sala de atendimento, cada canto foi montado para que voce "
        "se sinta bem enquanto cuida de si.\n\n"
        "Venha conhecer. WhatsApp (61) 99578-9867\n\n"
        "#unicorpos #planaltinadf #valedoamanhecer #clinicaespecializada"
    ),
    # Texto exato que foi ao ar em 18/08. A planilha guarda uma versao sem a
    # linha do RT; aqui fica o registro do que a conta realmente publicou.
    "F9_porta_fisio": (
        "Fisioterapia e quiropraxia na unidade da DF-130. 🧑‍⚕️\n\n"
        "A avaliação inicial define seu plano de sessões — nada começa sem ela.\n\n"
        "Sentido Vale do Amanhecer, com estacionamento no local.\n"
        "WhatsApp (61) 99578-9867\n\n"
        "Responsável Técnico: Mario Sergio Fernandes de Lima — CREFITO-11 nº 442563-F\n\n"
        "#unicorpos #fisioterapia #quiropraxia #planaltinadf #valedoamanhecer"
    ),
}

# ---------------------------------------------------------------------------
# Dias de Stories que dependem de conteudo humano (depoimento, enquete, video).
# Decisao do Leonardo: substituir por peca do banco de reserva.
# Ordem definida aqui, de proposito: sem estado, sem sorteio, sem surpresa.
# ---------------------------------------------------------------------------
SUBSTITUICAO = {
    6: "01_capa",
    9: "02_nova_unidade",
    12: "06_massagens",
    15: "07_fisioterapia",
    18: "15_lgpd",
    # O banco de reserva acaba no dia 18. Os tres dias abaixo usam pecas novas,
    # criadas por ferramentas/gerar_pecas.py, que seguem o tema que a planilha
    # pedia para o story daquele dia.
    21: "17_pos_laser",        # planilha: Stories "Estrutura"  -> virou dica pos-laser
    24: "18_primeira_fisio",   # planilha: Stories "Enquete: ja fez fisioterapia?"
    27: "19_biosseguranca",    # planilha: Stories "Bastidores: higienizacao"
}

# ---------------------------------------------------------------------------
# Conferido no perfil @unicorposclinica em 26/08/2026. O robo so entrou no ar
# em 26/08, entao tudo que foi ao ar antes disso foi publicado a mao.
# ---------------------------------------------------------------------------
PUBLICADO_EM = {
    1: "2026-08-13",   # F11_fachada
    2: "2026-08-14",   # F7_recepcao
    4: "2026-08-18",   # F9_porta_fisio (arte final diferente da prevista)
}

JA_PUBLICADO = set(PUBLICADO_EM)

# Dias uteis que passaram sem post porque o robo ainda nao estava ligado. As
# pecas nao foram queimadas: voltam para o banco de reserva do mes 2.
NAO_PUBLICADO = {
    3: "Dia util perdido: o robo ainda nao estava no ar e ninguem publicou a mao.",
    5: ("Dia util perdido. Em 19/08 foi publicado, no lugar, um carrossel sobre "
        "quantas sessoes de laser sao necessarias — peca criada fora do calendario."),
    6: "Dia util perdido: o robo ainda nao estava no ar e ninguem publicou a mao.",
    7: ("Dia util perdido. Era peca de odontologia, que so vai ao ar apos sua "
        "aprovacao por e-mail — e o SMTP ainda nao estava configurado."),
    8: "Dia util perdido: o robo ainda nao estava no ar e ninguem publicou a mao.",
    9: "Dia util perdido: o robo ainda nao estava no ar e ninguem publicou a mao.",
}


def limpar(txt):
    """Tira o marcador de publicado e espaco sobrando do nome do arquivo."""
    return re.sub(r"\s*✅.*$", "", str(txt or "")).strip()


def main():
    planilha = sys.argv[1] if len(sys.argv) > 1 else PLANILHA_PADRAO
    if not os.path.exists(planilha):
        sys.exit("Planilha nao encontrada: %s" % planilha)

    wb = openpyxl.load_workbook(planilha)

    # --- legendas -----------------------------------------------------------
    legendas = {}
    for linha in wb["Legendas prontas"].iter_rows(values_only=True):
        if linha and linha[0] and linha[1] and linha[0] not in ("Post",):
            legendas[limpar(linha[0])] = str(linha[1]).strip()
    legendas.update(LEGENDAS_EXTRA)

    # --- calendario ---------------------------------------------------------
    dias = {}
    for linha in wb["Calendario 30 dias"].iter_rows(values_only=True):
        if not linha or linha[0] is None:
            continue
        try:
            n = int(linha[0])
        except (TypeError, ValueError):
            continue

        formato = limpar(linha[1])
        arquivo = limpar(linha[2])
        publicado = "✅" in str(linha[2] or "")

        # A linha de Post manda. Se ja existe um Post para o dia, ignora a de Stories.
        if n in dias and dias[n].get("formato") == "Post":
            continue

        dias[n] = {
            "dia": n,
            "formato": formato,
            "arquivo_planilha": arquivo,
            "ja_publicado": publicado or n in JA_PUBLICADO,
        }

    # --- resolver a peca final de cada dia ----------------------------------
    saida = []
    for n in sorted(dias):
        d = dias[n]
        peca = d["arquivo_planilha"]
        substituido = False
        motivo = ""

        if n == 30 or d["formato"] == "Balanço":
            saida.append({
                "dia": n,
                "acao": "balanco",
                "observacao": "Fim do mes 1. Revisar alcance, salvamentos e conversas iniciadas.",
            })
            continue

        # Dia de Stories sem peca pronta -> banco de reserva
        if d["formato"] == "Stories" and peca not in legendas:
            if n in SUBSTITUICAO:
                peca = SUBSTITUICAO[n]
                substituido = True
                motivo = "Dia de story com conteudo humano (%s). Substituido por peca do banco de reserva." % d["arquivo_planilha"]
            else:
                saida.append({
                    "dia": n,
                    "acao": "avisar",
                    "observacao": "Story de '%s' precisa de conteudo seu e o banco de reserva acabou. Escolha uma peca e edite o calendario.json." % d["arquivo_planilha"],
                })
                continue

        categoria = CATEGORIA.get(peca, "institucional")
        if n in NAO_PUBLICADO:
            acao = "nao_publicado"
        elif d["ja_publicado"]:
            acao = "publicado_manualmente"
        elif categoria == "odonto":
            acao = "aprovar"
        else:
            acao = "publicar"
        registro = {
            "dia": n,
            "acao": acao,
            "arquivo": peca,
            "imagem": "img/%s.jpg" % peca,
            "categoria": categoria,
            "legenda": legendas.get(peca, ""),
            "alt_text": ALT.get(peca, ""),
        }
        if substituido:
            registro["substituido"] = True
            registro["observacao"] = motivo
        if categoria == "odonto" and acao != "nao_publicado":
            registro["observacao"] = (
                "Odontologia: nao publica sozinho. Vai por e-mail para aprovacao "
                "(Codigo de Etica Odontologica / CFO)."
            )
        if not registro["legenda"] and acao not in ("publicado_manualmente", "nao_publicado"):
            registro["acao"] = "avisar"
            registro["observacao"] = "Sem legenda na planilha para '%s'." % peca
        if n in PUBLICADO_EM:
            registro["publicado_em"] = PUBLICADO_EM[n]
        if n in NAO_PUBLICADO:
            registro["observacao"] = NAO_PUBLICADO[n]
        saida.append(registro)

    doc = {
        "_leia_me": "Gerado por ferramentas/gerar_calendario.py. Pode editar a mao: o robo le este arquivo, nao a planilha.",
        "dia_1": "2026-08-13",
        "fuso": "America/Sao_Paulo",
        "conta": "@unicorposclinica",
        "regras": [
            "Localizacao sempre 'Planaltina (Distrito Federal)' — nunca 'Planaltina de Goias', que e outra cidade.",
            "Texto alternativo sempre preenchido.",
            "Odontologia: nunca publicar tabela de precos (CFO).",
            "Publicidade odontologica cita a RT: Dra. Mayra Gabriela Alves Cardona — CRO-DF n CD-10122.",
            "Fisioterapia: RT Mario Sergio Fernandes de Lima — CREFITO-11 n 442563-F.",
        ],
        "banco_reserva_mes_2": sorted(
            d["arquivo"] for d in saida
            if d.get("acao") == "nao_publicado" and d.get("arquivo")
        ),
        "historico": [
            "26/08/2026 — perfil conferido: 4 posts no ar (13, 14, 18 e 19/08), todos "
            "publicados a mao. O robo nunca publicou: a unica execucao ate aqui foi um "
            "teste com dry_run marcado, que o GitHub mostrou como sucesso.",
            "Dias 3, 5, 6, 7, 8 e 9 passaram sem post. As pecas nao foram usadas e "
            "estao em banco_reserva_mes_2.",
            "A numeracao segue ancorada em dia_1 + dias uteis. Dia perdido nao empurra "
            "o resto do calendario: cada data continua com a peca que sempre teve.",
        ],
        "dias": saida,
    }

    destino = os.path.join(AQUI, "calendario.json")
    with open(destino, "w", encoding="utf-8") as fh:
        json.dump(doc, fh, ensure_ascii=False, indent=2)

    print("Escrito: %s" % destino)
    for d in saida:
        print("  Dia %2d  %-22s  %s" % (
            d["dia"], d.get("arquivo", "—"), d["acao"]))


if __name__ == "__main__":
    main()
